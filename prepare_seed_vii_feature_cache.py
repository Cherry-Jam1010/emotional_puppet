import argparse
import csv
import io
import json
import zipfile
from pathlib import Path
from typing import Dict, List

import numpy as np
import scipy.io as sio
from openpyxl import load_workbook

from seed_vii_data import EMOTION_TO_ID


ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_ZIP_PATH = ROOT_DIR / "SEED-VII.zip"
DEFAULT_LABEL_FILE = ROOT_DIR / "emotion_label_and_stimuli_order.xlsx"
DEFAULT_OUTPUT_DIR = ROOT_DIR / "artifacts" / "seed_vii" / "feature_cache"


def load_trial_labels(label_file: Path) -> Dict[int, str]:
    workbook = load_workbook(label_file, data_only=True)
    sheet = workbook.active

    labels: Dict[int, str] = {}
    trial_id = 1
    for row in sheet.iter_rows(values_only=True):
        header = row[0]
        if not isinstance(header, str) or not header.startswith("Session"):
            continue
        for emotion in row[1:]:
            if emotion is None:
                continue
            labels[trial_id] = str(emotion)
            trial_id += 1

    if len(labels) != 80:
        raise ValueError(f"Expected 80 labels, got {len(labels)}")
    return labels


def save_manifest(path: Path, rows: List[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "subject_id",
        "session_id",
        "trial_id",
        "emotion",
        "label_id",
        "seq_len",
        "feature_dim",
        "feature_path",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Cache SEED-VII EEG feature trials to .npy files.")
    parser.add_argument("--zip-path", type=Path, default=DEFAULT_ZIP_PATH)
    parser.add_argument("--label-file", type=Path, default=DEFAULT_LABEL_FILE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--feature-prefix", type=str, default="de_LDS")
    args = parser.parse_args()

    if not args.zip_path.exists():
        raise FileNotFoundError(f"Zip file not found: {args.zip_path}")
    if not args.label_file.exists():
        raise FileNotFoundError(f"Label file not found: {args.label_file}")

    trial_labels = load_trial_labels(args.label_file)
    feature_dir = args.output_dir / args.feature_prefix
    feature_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows: List[Dict[str, object]] = []
    subject_count = 0

    with zipfile.ZipFile(args.zip_path, "r") as zf:
        feature_files = sorted(
            [name for name in zf.namelist() if name.startswith("SEED-VII/EEG_features/") and name.endswith(".mat")],
            key=lambda name: int(Path(name).stem),
        )

        for zip_name in feature_files:
            subject_id = int(Path(zip_name).stem)
            subject_count += 1
            subject_dir = feature_dir / f"subject_{subject_id:02d}"
            subject_dir.mkdir(parents=True, exist_ok=True)

            with zf.open(zip_name) as f:
                mat_data = sio.loadmat(io.BytesIO(f.read()))

            for trial_id in range(1, 81):
                key = f"{args.feature_prefix}_{trial_id}"
                if key not in mat_data:
                    raise KeyError(f"Missing key {key} in {zip_name}")

                feature = np.asarray(mat_data[key], dtype=np.float32).reshape(mat_data[key].shape[0], -1)
                output_path = subject_dir / f"trial_{trial_id:02d}.npy"
                np.save(output_path, feature)

                emotion = trial_labels[trial_id]
                manifest_rows.append(
                    {
                        "subject_id": subject_id,
                        "session_id": (trial_id - 1) // 20 + 1,
                        "trial_id": trial_id,
                        "emotion": emotion,
                        "label_id": EMOTION_TO_ID[emotion],
                        "seq_len": int(feature.shape[0]),
                        "feature_dim": int(feature.shape[1]),
                        "feature_path": str(output_path.resolve()),
                    }
                )

    manifest_path = feature_dir / "manifest.csv"
    summary_path = feature_dir / "summary.json"
    save_manifest(manifest_path, manifest_rows)
    summary = {
        "feature_prefix": args.feature_prefix,
        "num_subjects": subject_count,
        "num_trials": len(manifest_rows),
        "output_dir": str(feature_dir.resolve()),
        "manifest_path": str(manifest_path.resolve()),
        "feature_dim": int(manifest_rows[0]["feature_dim"]) if manifest_rows else 0,
    }
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(f"Cached {len(manifest_rows)} trials to: {feature_dir}")
    print(f"Manifest written to: {manifest_path}")
    print(f"Summary written to: {summary_path}")


if __name__ == "__main__":
    main()
