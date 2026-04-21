import argparse
import csv
import io
import json
import random
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import numpy as np
import torch
import torch.nn as nn
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, classification_report, f1_score
from torch.utils.data import DataLoader, Dataset

from seed_vii_models import ConvFeatureClassifier
from seed_vii_data import EMOTION_TO_ID


ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_ZIP_PATH = ROOT_DIR / "SEED-VII.zip"
DEFAULT_LABEL_FILE = ROOT_DIR / "emotion_label_and_stimuli_order.xlsx"
DEFAULT_RUN_DIR = ROOT_DIR / "artifacts" / "seed_vii" / "feature_window_training"
ID_TO_EMOTION = {value: key for key, value in EMOTION_TO_ID.items()}


def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)


def load_trial_labels(label_file: Path) -> Dict[int, int]:
    workbook = load_workbook(label_file, data_only=True)
    sheet = workbook.active
    labels: Dict[int, int] = {}
    trial_id = 1
    for row in sheet.iter_rows(values_only=True):
        header = row[0]
        if not isinstance(header, str) or not header.startswith("Session"):
            continue
        for emotion in row[1:]:
            if emotion is None:
                continue
            labels[trial_id] = EMOTION_TO_ID[str(emotion)]
            trial_id += 1
    return labels


def load_feature_trials(zip_path: Path, label_map: Dict[int, int], feature_prefix: str) -> List[Dict[str, object]]:
    import scipy.io as sio

    records: List[Dict[str, object]] = []
    with zipfile.ZipFile(zip_path, "r") as zf:
        feature_files = sorted(
            [name for name in zf.namelist() if name.startswith("SEED-VII/EEG_features/") and name.endswith(".mat")],
            key=lambda name: int(Path(name).stem),
        )
        for name in feature_files:
            subject_id = int(Path(name).stem)
            with zf.open(name) as f:
                mat_data = sio.loadmat(io.BytesIO(f.read()))
            for trial_id in range(1, 81):
                key = f"{feature_prefix}_{trial_id}"
                feature = np.asarray(mat_data[key], dtype=np.float32).reshape(mat_data[key].shape[0], -1)
                records.append(
                    {
                        "subject_id": subject_id,
                        "session_id": (trial_id - 1) // 20 + 1,
                        "trial_id": trial_id,
                        "label_id": label_map[trial_id],
                        "emotion": ID_TO_EMOTION[label_map[trial_id]],
                        "feature": feature,
                        "seq_len": int(feature.shape[0]),
                        "feature_dim": int(feature.shape[1]),
                    }
                )
    return records


def load_feature_trials_from_manifest(manifest_path: Path) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    with manifest_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            feature_path = Path(row["feature_path"])
            feature = np.load(feature_path).astype(np.float32)
            records.append(
                {
                    "subject_id": int(row["subject_id"]),
                    "session_id": int(row["session_id"]),
                    "trial_id": int(row["trial_id"]),
                    "label_id": int(row["label_id"]),
                    "emotion": row["emotion"],
                    "feature": feature,
                    "seq_len": int(row["seq_len"]),
                    "feature_dim": int(row["feature_dim"]),
                }
            )
    return records


def split_subjects(
    subjects: Sequence[int],
    train_ratio: float,
    val_ratio: float,
    seed: int,
) -> Tuple[List[int], List[int], List[int]]:
    rng = random.Random(seed)
    subject_list = sorted(subjects)
    rng.shuffle(subject_list)
    n_subjects = len(subject_list)
    n_train = max(1, int(round(n_subjects * train_ratio)))
    n_val = max(1, int(round(n_subjects * val_ratio)))
    n_test = n_subjects - n_train - n_val
    if n_test < 1:
        n_test = 1
        if n_train > n_val and n_train > 1:
            n_train -= 1
        else:
            n_val -= 1
    return (
        sorted(subject_list[:n_train]),
        sorted(subject_list[n_train:n_train + n_val]),
        sorted(subject_list[n_train + n_val:]),
    )


def filter_records(records: Sequence[Dict[str, object]], subjects: Sequence[int]) -> List[Dict[str, object]]:
    subject_set = set(subjects)
    return [record for record in records if int(record["subject_id"]) in subject_set]


def build_window_indices(
    records: Sequence[Dict[str, object]],
    window_len: int,
    stride: int,
    evenly_spaced_limit: int = 0,
) -> List[Dict[str, object]]:
    samples: List[Dict[str, object]] = []
    for record in records:
        seq_len = int(record["seq_len"])
        if seq_len < window_len:
            continue
        starts = list(range(0, seq_len - window_len + 1, stride))
        if evenly_spaced_limit > 0 and len(starts) > evenly_spaced_limit:
            positions = np.linspace(0, len(starts) - 1, evenly_spaced_limit)
            starts = [starts[int(round(pos))] for pos in positions]
            starts = sorted(set(starts))
        for start in starts:
            samples.append(
                {
                    "feature": record["feature"],
                    "label_id": int(record["label_id"]),
                    "subject_id": int(record["subject_id"]),
                    "trial_id": int(record["trial_id"]),
                    "start": start,
                    "stop": start + window_len,
                }
            )
    return samples


class FeatureWindowDataset(Dataset):
    def __init__(self, samples: Sequence[Dict[str, object]], normalize: bool = True) -> None:
        self.samples = list(samples)
        self.normalize = normalize

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, index: int):
        sample = self.samples[index]
        feature = np.asarray(sample["feature"][sample["start"]:sample["stop"]], dtype=np.float32)
        if self.normalize:
            mean = feature.mean(axis=0, keepdims=True)
            std = feature.std(axis=0, keepdims=True)
            feature = (feature - mean) / np.maximum(std, 1e-6)
        x = torch.from_numpy(feature.transpose(1, 0))
        return (
            x,
            torch.tensor(sample["label_id"], dtype=torch.long),
            torch.tensor(sample["subject_id"], dtype=torch.long),
            torch.tensor(sample["trial_id"], dtype=torch.long),
        )


def build_class_weights(records: Sequence[Dict[str, object]]) -> torch.Tensor:
    counts = Counter(int(record["label_id"]) for record in records)
    total = sum(counts.values())
    n_classes = len(EMOTION_TO_ID)
    weights = [total / (n_classes * counts[class_id]) for class_id in range(n_classes)]
    return torch.tensor(weights, dtype=torch.float32)


def aggregate_trial_predictions(
    logits_list: List[np.ndarray],
    labels_list: List[int],
    subject_ids: List[int],
    trial_ids: List[int],
) -> Tuple[np.ndarray, np.ndarray]:
    grouped_logits: Dict[Tuple[int, int], List[np.ndarray]] = defaultdict(list)
    grouped_labels: Dict[Tuple[int, int], int] = {}
    for logits, label, subject_id, trial_id in zip(logits_list, labels_list, subject_ids, trial_ids):
        key = (subject_id, trial_id)
        grouped_logits[key].append(logits)
        grouped_labels[key] = label

    y_true: List[int] = []
    y_pred: List[int] = []
    for key in sorted(grouped_logits):
        mean_logits = np.mean(grouped_logits[key], axis=0)
        y_true.append(grouped_labels[key])
        y_pred.append(int(np.argmax(mean_logits)))
    return np.asarray(y_true), np.asarray(y_pred)


def train_one_epoch(
    model: nn.Module,
    loader: DataLoader,
    optimizer: torch.optim.Optimizer,
    criterion: nn.Module,
    device: torch.device,
) -> float:
    model.train()
    total_loss = 0.0
    total_examples = 0
    for x, y, _, _ in loader:
        x = x.to(device)
        y = y.to(device)
        optimizer.zero_grad(set_to_none=True)
        logits = model(x)
        loss = criterion(logits, y)
        loss.backward()
        optimizer.step()
        batch_size = y.size(0)
        total_loss += loss.item() * batch_size
        total_examples += batch_size
    return total_loss / max(total_examples, 1)


def evaluate(model: nn.Module, loader: DataLoader, criterion: nn.Module, device: torch.device) -> Dict[str, object]:
    model.eval()
    total_loss = 0.0
    total_examples = 0
    logits_list: List[np.ndarray] = []
    labels_list: List[int] = []
    subject_ids: List[int] = []
    trial_ids: List[int] = []
    with torch.no_grad():
        for x, y, batch_subject_ids, batch_trial_ids in loader:
            x = x.to(device)
            y = y.to(device)
            logits = model(x)
            loss = criterion(logits, y)
            batch_size = y.size(0)
            total_loss += loss.item() * batch_size
            total_examples += batch_size
            logits_list.extend(logits.cpu().numpy())
            labels_list.extend(y.cpu().numpy().tolist())
            subject_ids.extend(batch_subject_ids.numpy().tolist())
            trial_ids.extend(batch_trial_ids.numpy().tolist())

    y_true, y_pred = aggregate_trial_predictions(logits_list, labels_list, subject_ids, trial_ids)
    report = classification_report(
        y_true,
        y_pred,
        target_names=[ID_TO_EMOTION[i] for i in range(len(ID_TO_EMOTION))],
        output_dict=True,
        zero_division=0,
    )
    return {
        "loss": total_loss / max(total_examples, 1),
        "accuracy": accuracy_score(y_true, y_pred),
        "macro_f1": f1_score(y_true, y_pred, average="macro"),
        "report": report,
        "num_trials": len(y_true),
    }


def save_json(path: Path, payload: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Train window-based SEED-VII feature classifier.")
    parser.add_argument("--zip-path", type=Path, default=DEFAULT_ZIP_PATH)
    parser.add_argument("--label-file", type=Path, default=DEFAULT_LABEL_FILE)
    parser.add_argument("--cache-manifest", type=Path, default=None)
    parser.add_argument("--run-dir", type=Path, default=DEFAULT_RUN_DIR)
    parser.add_argument("--feature-prefix", type=str, default="de_LDS")
    parser.add_argument("--window-len", type=int, default=8)
    parser.add_argument("--train-stride", type=int, default=4)
    parser.add_argument("--eval-stride", type=int, default=4)
    parser.add_argument("--eval-max-windows", type=int, default=12)
    parser.add_argument("--train-ratio", type=float, default=0.7)
    parser.add_argument("--val-ratio", type=float, default=0.15)
    parser.add_argument("--batch-size", type=int, default=64)
    parser.add_argument("--epochs", type=int, default=15)
    parser.add_argument("--dropout", type=float, default=0.25)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--weight-decay", type=float, default=1e-4)
    parser.add_argument("--patience", type=int, default=4)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    set_seed(args.seed)
    device = torch.device("cpu")

    if args.cache_manifest is not None:
        if not args.cache_manifest.exists():
            raise FileNotFoundError(f"Cache manifest not found: {args.cache_manifest}")
        records = load_feature_trials_from_manifest(args.cache_manifest)
    else:
        label_map = load_trial_labels(args.label_file)
        records = load_feature_trials(args.zip_path, label_map, args.feature_prefix)
    input_dim = int(records[0]["feature_dim"])

    unique_subjects = sorted({int(record["subject_id"]) for record in records})
    train_subjects, val_subjects, test_subjects = split_subjects(unique_subjects, args.train_ratio, args.val_ratio, args.seed)
    train_records = filter_records(records, train_subjects)
    val_records = filter_records(records, val_subjects)
    test_records = filter_records(records, test_subjects)

    train_samples = build_window_indices(train_records, args.window_len, args.train_stride)
    val_samples = build_window_indices(val_records, args.window_len, args.eval_stride, evenly_spaced_limit=args.eval_max_windows)
    test_samples = build_window_indices(test_records, args.window_len, args.eval_stride, evenly_spaced_limit=args.eval_max_windows)

    train_dataset = FeatureWindowDataset(train_samples)
    val_dataset = FeatureWindowDataset(val_samples)
    test_dataset = FeatureWindowDataset(test_samples)

    train_loader = DataLoader(train_dataset, batch_size=args.batch_size, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=args.batch_size, shuffle=False)
    test_loader = DataLoader(test_dataset, batch_size=args.batch_size, shuffle=False)

    model = ConvFeatureClassifier(input_dim=input_dim, n_classes=len(EMOTION_TO_ID), dropout=args.dropout).to(device)
    optimizer = torch.optim.AdamW(model.parameters(), lr=args.lr, weight_decay=args.weight_decay)
    class_weights = build_class_weights(train_samples).to(device)
    criterion = nn.CrossEntropyLoss(weight=class_weights, label_smoothing=0.02)

    args.run_dir.mkdir(parents=True, exist_ok=True)
    checkpoint_path = args.run_dir / "best_model.pt"
    history_path = args.run_dir / "history.json"
    results_path = args.run_dir / "results.json"
    split_path = args.run_dir / "subject_split.json"

    save_json(
        split_path,
        {
            "train_subjects": train_subjects,
            "val_subjects": val_subjects,
            "test_subjects": test_subjects,
        },
    )

    best_val_f1 = -1.0
    best_epoch = -1
    stale_epochs = 0
    history: List[Dict[str, float]] = []

    for epoch in range(1, args.epochs + 1):
        train_loss = train_one_epoch(model, train_loader, optimizer, criterion, device)
        val_metrics = evaluate(model, val_loader, criterion, device)
        history.append(
            {
                "epoch": epoch,
                "train_loss": float(train_loss),
                "val_loss": float(val_metrics["loss"]),
                "val_accuracy": float(val_metrics["accuracy"]),
                "val_macro_f1": float(val_metrics["macro_f1"]),
            }
        )
        print(
            f"Epoch {epoch:02d} | train_loss={train_loss:.4f} | "
            f"val_loss={val_metrics['loss']:.4f} | "
            f"val_acc={val_metrics['accuracy']:.4f} | "
            f"val_macro_f1={val_metrics['macro_f1']:.4f}"
        )

        if val_metrics["macro_f1"] > best_val_f1:
            best_val_f1 = float(val_metrics["macro_f1"])
            best_epoch = epoch
            stale_epochs = 0
            torch.save(
                {
                    "model_state_dict": model.state_dict(),
                    "input_dim": input_dim,
                    "window_len": args.window_len,
                    "feature_prefix": args.feature_prefix,
                    "best_epoch": best_epoch,
                    "best_val_macro_f1": best_val_f1,
                },
                checkpoint_path,
            )
        else:
            stale_epochs += 1
            if stale_epochs >= args.patience:
                print(f"Early stopping at epoch {epoch}.")
                break

    save_json(history_path, {"history": history})

    checkpoint = torch.load(checkpoint_path, map_location=device, weights_only=True)
    model.load_state_dict(checkpoint["model_state_dict"])

    val_metrics = evaluate(model, val_loader, criterion, device)
    test_metrics = evaluate(model, test_loader, criterion, device)
    results = {
        "best_epoch": best_epoch,
        "best_val_macro_f1": best_val_f1,
        "feature_prefix": args.feature_prefix,
        "window_len": args.window_len,
        "dataset_sizes": {
            "train_windows": len(train_samples),
            "val_windows": len(val_samples),
            "test_windows": len(test_samples),
            "train_trials": len(train_records),
            "val_trials": len(val_records),
            "test_trials": len(test_records),
        },
        "val_metrics": {
            "loss": float(val_metrics["loss"]),
            "accuracy": float(val_metrics["accuracy"]),
            "macro_f1": float(val_metrics["macro_f1"]),
            "num_trials": int(val_metrics["num_trials"]),
            "report": val_metrics["report"],
        },
        "test_metrics": {
            "loss": float(test_metrics["loss"]),
            "accuracy": float(test_metrics["accuracy"]),
            "macro_f1": float(test_metrics["macro_f1"]),
            "num_trials": int(test_metrics["num_trials"]),
            "report": test_metrics["report"],
        },
    }
    save_json(results_path, results)

    print(f"Best checkpoint: {checkpoint_path}")
    print(f"Results written to: {results_path}")
    print(
        f"Test accuracy={results['test_metrics']['accuracy']:.4f}, "
        f"test macro_f1={results['test_metrics']['macro_f1']:.4f}"
    )


if __name__ == "__main__":
    main()
