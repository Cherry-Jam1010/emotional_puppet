import argparse
import io
import json
import random
import zipfile
from collections import Counter
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import numpy as np
import torch
import torch.nn as nn
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, classification_report, f1_score
from torch.nn.utils.rnn import pack_padded_sequence, pad_packed_sequence, pad_sequence
from torch.utils.data import DataLoader, Dataset

from seed_vii_data import EMOTION_TO_ID


ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_ZIP_PATH = ROOT_DIR / "SEED-VII.zip"
DEFAULT_LABEL_FILE = ROOT_DIR / "emotion_label_and_stimuli_order.xlsx"
DEFAULT_RUN_DIR = ROOT_DIR / "artifacts" / "seed_vii" / "feature_training"
ID_TO_EMOTION = {value: key for key, value in EMOTION_TO_ID.items()}


def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)


def load_trial_labels(label_file: Path) -> Dict[int, int]:
    workbook = load_workbook(label_file, data_only=True)
    sheet = workbook.active
    labels: Dict[int, int] = {}
    trial_id = 1

    for row in sheet.iter_rows(values_only=True):
        header = row[0]
        if not isinstance(header, str) or not header.startswith("Session"):
            continue
        for emotion in row[1:]:
            if emotion is None:
                continue
            labels[trial_id] = EMOTION_TO_ID[str(emotion)]
            trial_id += 1

    if len(labels) != 80:
        raise ValueError(f"Expected 80 labels, got {len(labels)}")
    return labels


def load_feature_trials(
    zip_path: Path,
    label_map: Dict[int, int],
    feature_prefix: str,
) -> List[Dict[str, object]]:
    import scipy.io as sio

    records: List[Dict[str, object]] = []
    with zipfile.ZipFile(zip_path, "r") as zf:
        feature_files = sorted(
            [name for name in zf.namelist() if name.startswith("SEED-VII/EEG_features/") and name.endswith(".mat")],
            key=lambda name: int(Path(name).stem),
        )

        for name in feature_files:
            subject_id = int(Path(name).stem)
            with zf.open(name) as f:
                mat_data = sio.loadmat(io.BytesIO(f.read()))

            for trial_id in range(1, 81):
                key = f"{feature_prefix}_{trial_id}"
                if key not in mat_data:
                    raise KeyError(f"Missing key {key} in {name}")

                feature = np.asarray(mat_data[key], dtype=np.float32)
                feature = feature.reshape(feature.shape[0], -1)
                session_id = (trial_id - 1) // 20 + 1

                records.append(
                    {
                        "subject_id": subject_id,
                        "session_id": session_id,
                        "trial_id": trial_id,
                        "label_id": label_map[trial_id],
                        "emotion": ID_TO_EMOTION[label_map[trial_id]],
                        "feature": feature,
                        "seq_len": int(feature.shape[0]),
                        "feature_dim": int(feature.shape[1]),
                    }
                )

    return records


def split_subjects(
    subjects: Sequence[int],
    train_ratio: float,
    val_ratio: float,
    seed: int,
) -> Tuple[List[int], List[int], List[int]]:
    rng = random.Random(seed)
    subject_list = sorted(subjects)
    rng.shuffle(subject_list)

    n_subjects = len(subject_list)
    n_train = max(1, int(round(n_subjects * train_ratio)))
    n_val = max(1, int(round(n_subjects * val_ratio)))
    n_test = n_subjects - n_train - n_val

    if n_test < 1:
        n_test = 1
        if n_train > n_val and n_train > 1:
            n_train -= 1
        else:
            n_val -= 1

    train_subjects = sorted(subject_list[:n_train])
    val_subjects = sorted(subject_list[n_train:n_train + n_val])
    test_subjects = sorted(subject_list[n_train + n_val:])
    return train_subjects, val_subjects, test_subjects


def filter_records(records: Sequence[Dict[str, object]], subjects: Sequence[int]) -> List[Dict[str, object]]:
    subject_set = set(subjects)
    return [record for record in records if int(record["subject_id"]) in subject_set]


class SeedVIIFeatureDataset(Dataset):
    def __init__(self, records: Sequence[Dict[str, object]], normalize: bool = True) -> None:
        self.records = list(records)
        self.normalize = normalize

    def __len__(self) -> int:
        return len(self.records)

    def __getitem__(self, index: int):
        record = self.records[index]
        feature = np.asarray(record["feature"], dtype=np.float32)

        if self.normalize:
            mean = feature.mean(axis=0, keepdims=True)
            std = feature.std(axis=0, keepdims=True)
            feature = (feature - mean) / np.maximum(std, 1e-6)

        return {
            "x": torch.from_numpy(feature),
            "y": int(record["label_id"]),
            "subject_id": int(record["subject_id"]),
            "trial_id": int(record["trial_id"]),
            "seq_len": int(record["seq_len"]),
        }


def collate_batch(batch: Sequence[Dict[str, object]]) -> Dict[str, torch.Tensor]:
    features = [item["x"] for item in batch]
    lengths = torch.tensor([item["seq_len"] for item in batch], dtype=torch.long)
    labels = torch.tensor([item["y"] for item in batch], dtype=torch.long)
    subject_ids = torch.tensor([item["subject_id"] for item in batch], dtype=torch.long)
    trial_ids = torch.tensor([item["trial_id"] for item in batch], dtype=torch.long)
    padded = pad_sequence(features, batch_first=True)
    return {
        "x": padded,
        "lengths": lengths,
        "y": labels,
        "subject_id": subject_ids,
        "trial_id": trial_ids,
    }


class GRUEmotionClassifier(nn.Module):
    def __init__(
        self,
        input_dim: int,
        hidden_dim: int,
        n_layers: int,
        n_classes: int,
        dropout: float,
    ) -> None:
        super().__init__()
        self.input_proj = nn.Sequential(
            nn.LayerNorm(input_dim),
            nn.Linear(input_dim, hidden_dim),
            nn.GELU(),
            nn.Dropout(dropout),
        )
        self.encoder = nn.GRU(
            input_size=hidden_dim,
            hidden_size=hidden_dim,
            num_layers=n_layers,
            batch_first=True,
            dropout=dropout if n_layers > 1 else 0.0,
            bidirectional=True,
        )
        self.classifier = nn.Sequential(
            nn.Linear(hidden_dim * 2, hidden_dim),
            nn.GELU(),
            nn.Dropout(dropout),
            nn.Linear(hidden_dim, n_classes),
        )

    def forward(self, x: torch.Tensor, lengths: torch.Tensor) -> torch.Tensor:
        x = self.input_proj(x)
        packed = pack_padded_sequence(x, lengths.cpu(), batch_first=True, enforce_sorted=False)
        packed_output, _ = self.encoder(packed)
        output, _ = pad_packed_sequence(packed_output, batch_first=True)

        mask = torch.arange(output.size(1), device=output.device).unsqueeze(0) < lengths.unsqueeze(1)
        mask = mask.unsqueeze(-1)
        summed = (output * mask).sum(dim=1)
        pooled = summed / lengths.unsqueeze(1).clamp_min(1)
        return self.classifier(pooled)


def build_class_weights(records: Sequence[Dict[str, object]]) -> torch.Tensor:
    counts = Counter(int(record["label_id"]) for record in records)
    total = sum(counts.values())
    weights = []
    n_classes = len(EMOTION_TO_ID)
    for class_id in range(n_classes):
        weights.append(total / (n_classes * counts[class_id]))
    return torch.tensor(weights, dtype=torch.float32)


def train_one_epoch(
    model: nn.Module,
    loader: DataLoader,
    optimizer: torch.optim.Optimizer,
    criterion: nn.Module,
    device: torch.device,
) -> float:
    model.train()
    total_loss = 0.0
    total_examples = 0

    for batch in loader:
        x = batch["x"].to(device)
        lengths = batch["lengths"].to(device)
        y = batch["y"].to(device)

        optimizer.zero_grad(set_to_none=True)
        logits = model(x, lengths)
        loss = criterion(logits, y)
        loss.backward()
        optimizer.step()

        batch_size = y.size(0)
        total_loss += loss.item() * batch_size
        total_examples += batch_size

    return total_loss / max(total_examples, 1)


def evaluate(
    model: nn.Module,
    loader: DataLoader,
    criterion: nn.Module,
    device: torch.device,
) -> Dict[str, object]:
    model.eval()
    total_loss = 0.0
    total_examples = 0
    y_true: List[int] = []
    y_pred: List[int] = []

    with torch.no_grad():
        for batch in loader:
            x = batch["x"].to(device)
            lengths = batch["lengths"].to(device)
            y = batch["y"].to(device)

            logits = model(x, lengths)
            loss = criterion(logits, y)

            batch_size = y.size(0)
            total_loss += loss.item() * batch_size
            total_examples += batch_size

            pred = torch.argmax(logits, dim=1)
            y_true.extend(y.cpu().numpy().tolist())
            y_pred.extend(pred.cpu().numpy().tolist())

    report = classification_report(
        y_true,
        y_pred,
        target_names=[ID_TO_EMOTION[i] for i in range(len(ID_TO_EMOTION))],
        output_dict=True,
        zero_division=0,
    )

    return {
        "loss": total_loss / max(total_examples, 1),
        "accuracy": accuracy_score(y_true, y_pred),
        "macro_f1": f1_score(y_true, y_pred, average="macro"),
        "report": report,
        "num_trials": len(y_true),
    }


def save_json(path: Path, payload: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Train a fast baseline on SEED-VII EEG feature sequences.")
    parser.add_argument("--zip-path", type=Path, default=DEFAULT_ZIP_PATH)
    parser.add_argument("--label-file", type=Path, default=DEFAULT_LABEL_FILE)
    parser.add_argument("--run-dir", type=Path, default=DEFAULT_RUN_DIR)
    parser.add_argument("--feature-prefix", type=str, default="de_LDS")
    parser.add_argument("--train-ratio", type=float, default=0.7)
    parser.add_argument("--val-ratio", type=float, default=0.15)
    parser.add_argument("--batch-size", type=int, default=32)
    parser.add_argument("--epochs", type=int, default=20)
    parser.add_argument("--hidden-dim", type=int, default=128)
    parser.add_argument("--num-layers", type=int, default=2)
    parser.add_argument("--dropout", type=float, default=0.3)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--weight-decay", type=float, default=1e-4)
    parser.add_argument("--patience", type=int, default=5)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    if not args.zip_path.exists():
        raise FileNotFoundError(f"Zip file not found: {args.zip_path}")
    if not args.label_file.exists():
        raise FileNotFoundError(f"Label file not found: {args.label_file}")

    set_seed(args.seed)
    device = torch.device("cpu")

    label_map = load_trial_labels(args.label_file)
    records = load_feature_trials(args.zip_path, label_map, args.feature_prefix)
    unique_subjects = sorted({int(record["subject_id"]) for record in records})
    train_subjects, val_subjects, test_subjects = split_subjects(
        unique_subjects,
        args.train_ratio,
        args.val_ratio,
        args.seed,
    )

    train_records = filter_records(records, train_subjects)
    val_records = filter_records(records, val_subjects)
    test_records = filter_records(records, test_subjects)

    input_dim = int(train_records[0]["feature_dim"])
    train_dataset = SeedVIIFeatureDataset(train_records)
    val_dataset = SeedVIIFeatureDataset(val_records)
    test_dataset = SeedVIIFeatureDataset(test_records)

    train_loader = DataLoader(train_dataset, batch_size=args.batch_size, shuffle=True, collate_fn=collate_batch)
    val_loader = DataLoader(val_dataset, batch_size=args.batch_size, shuffle=False, collate_fn=collate_batch)
    test_loader = DataLoader(test_dataset, batch_size=args.batch_size, shuffle=False, collate_fn=collate_batch)

    model = GRUEmotionClassifier(
        input_dim=input_dim,
        hidden_dim=args.hidden_dim,
        n_layers=args.num_layers,
        n_classes=len(EMOTION_TO_ID),
        dropout=args.dropout,
    ).to(device)
    optimizer = torch.optim.AdamW(model.parameters(), lr=args.lr, weight_decay=args.weight_decay)
    class_weights = build_class_weights(train_records).to(device)
    criterion = nn.CrossEntropyLoss(weight=class_weights, label_smoothing=0.03)

    args.run_dir.mkdir(parents=True, exist_ok=True)
    checkpoint_path = args.run_dir / "best_model.pt"
    results_path = args.run_dir / "results.json"
    history_path = args.run_dir / "history.json"
    split_path = args.run_dir / "subject_split.json"

    save_json(
        split_path,
        {
            "train_subjects": train_subjects,
            "val_subjects": val_subjects,
            "test_subjects": test_subjects,
        },
    )

    history: List[Dict[str, float]] = []
    best_val_f1 = -1.0
    best_epoch = -1
    stale_epochs = 0

    for epoch in range(1, args.epochs + 1):
        train_loss = train_one_epoch(model, train_loader, optimizer, criterion, device)
        val_metrics = evaluate(model, val_loader, criterion, device)

        history.append(
            {
                "epoch": epoch,
                "train_loss": float(train_loss),
                "val_loss": float(val_metrics["loss"]),
                "val_accuracy": float(val_metrics["accuracy"]),
                "val_macro_f1": float(val_metrics["macro_f1"]),
            }
        )
        print(
            f"Epoch {epoch:02d} | train_loss={train_loss:.4f} | "
            f"val_loss={val_metrics['loss']:.4f} | "
            f"val_acc={val_metrics['accuracy']:.4f} | "
            f"val_macro_f1={val_metrics['macro_f1']:.4f}"
        )

        if val_metrics["macro_f1"] > best_val_f1:
            best_val_f1 = float(val_metrics["macro_f1"])
            best_epoch = epoch
            stale_epochs = 0
            torch.save(
                {
                    "model_state_dict": model.state_dict(),
                    "feature_prefix": args.feature_prefix,
                    "input_dim": input_dim,
                    "hidden_dim": args.hidden_dim,
                    "num_layers": args.num_layers,
                    "dropout": args.dropout,
                    "best_epoch": best_epoch,
                    "best_val_macro_f1": best_val_f1,
                },
                checkpoint_path,
            )
        else:
            stale_epochs += 1
            if stale_epochs >= args.patience:
                print(f"Early stopping at epoch {epoch}.")
                break

    save_json(history_path, {"history": history})

    checkpoint = torch.load(checkpoint_path, map_location=device)
    model.load_state_dict(checkpoint["model_state_dict"])

    val_metrics = evaluate(model, val_loader, criterion, device)
    test_metrics = evaluate(model, test_loader, criterion, device)
    results = {
        "best_epoch": best_epoch,
        "best_val_macro_f1": best_val_f1,
        "feature_prefix": args.feature_prefix,
        "dataset_sizes": {
            "train_trials": len(train_records),
            "val_trials": len(val_records),
            "test_trials": len(test_records),
        },
        "val_metrics": {
            "loss": float(val_metrics["loss"]),
            "accuracy": float(val_metrics["accuracy"]),
            "macro_f1": float(val_metrics["macro_f1"]),
            "num_trials": int(val_metrics["num_trials"]),
            "report": val_metrics["report"],
        },
        "test_metrics": {
            "loss": float(test_metrics["loss"]),
            "accuracy": float(test_metrics["accuracy"]),
            "macro_f1": float(test_metrics["macro_f1"]),
            "num_trials": int(test_metrics["num_trials"]),
            "report": test_metrics["report"],
        },
    }
    save_json(results_path, results)

    print(f"Best checkpoint: {checkpoint_path}")
    print(f"Results written to: {results_path}")
    print(
        f"Test accuracy={results['test_metrics']['accuracy']:.4f}, "
        f"test macro_f1={results['test_metrics']['macro_f1']:.4f}"
    )


if __name__ == "__main__":
    main()
