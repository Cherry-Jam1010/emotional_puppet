import argparse
import csv
import json
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import scipy.io as sio
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_EEG_DIR = ROOT_DIR / "SEED-VII_EEG_preprocessed" / "SEED-VII" / "EEG_preprocessed"
DEFAULT_LABEL_FILE = ROOT_DIR / "emotion_label_and_stimuli_order.xlsx"
DEFAULT_OUTPUT_DIR = ROOT_DIR / "artifacts" / "seed_vii"
DEFAULT_SAMPLING_RATE = 200

EMOTION_TO_ID = {
    "Anger": 0,
    "Disgust": 1,
    "Fear": 2,
    "Happy": 3,
    "Neutral": 4,
    "Sad": 5,
    "Surprise": 6,
}


def load_trial_labels(label_file: Path) -> Dict[int, str]:
    workbook = load_workbook(label_file, data_only=True)
    sheet = workbook.active

    trial_labels: Dict[int, str] = {}
    trial_index = 1

    for row in sheet.iter_rows(values_only=True):
        header = row[0]
        if not isinstance(header, str) or not header.startswith("Session"):
            continue

        for emotion in row[1:]:
            if emotion is None:
                continue
            trial_labels[trial_index] = str(emotion)
            trial_index += 1

    if len(trial_labels) != 80:
        raise ValueError(f"Expected 80 trial labels, got {len(trial_labels)}")

    return trial_labels


def load_subject_trials(mat_path: Path) -> Dict[int, np.ndarray]:
    mat_data = sio.loadmat(mat_path)
    trial_keys = sorted(
        [key for key in mat_data.keys() if not key.startswith("__")],
        key=int,
    )

    trials: Dict[int, np.ndarray] = {}
    for key in trial_keys:
        array = np.asarray(mat_data[key], dtype=np.float32)
        trials[int(key)] = array

    if len(trials) != 80:
        raise ValueError(f"{mat_path.name} should contain 80 trials, got {len(trials)}")

    return trials


def build_trial_records(
    eeg_dir: Path,
    trial_labels: Dict[int, str],
    sampling_rate: int,
) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []

    for mat_path in sorted(eeg_dir.glob("*.mat"), key=lambda path: int(path.stem)):
        subject_id = int(mat_path.stem)
        trials = load_subject_trials(mat_path)

        for trial_id, signal in trials.items():
            emotion = trial_labels[trial_id]
            label_id = EMOTION_TO_ID[emotion]
            session_id = (trial_id - 1) // 20 + 1
            n_channels, n_samples = signal.shape

            records.append(
                {
                    "subject_id": subject_id,
                    "session_id": session_id,
                    "trial_id": trial_id,
                    "emotion": emotion,
                    "label_id": label_id,
                    "n_channels": n_channels,
                    "n_samples": n_samples,
                    "duration_sec": round(n_samples / float(sampling_rate), 4),
                    "file_path": str(mat_path.resolve()),
                    "mat_key": str(trial_id),
                }
            )

    return records


def build_window_records(
    trial_records: List[Dict[str, object]],
    sampling_rate: int,
    window_seconds: float,
    stride_seconds: float,
) -> List[Dict[str, object]]:
    window_size = int(round(window_seconds * sampling_rate))
    stride_size = int(round(stride_seconds * sampling_rate))

    if window_size <= 0 or stride_size <= 0:
        raise ValueError("window_seconds and stride_seconds must be positive")

    window_records: List[Dict[str, object]] = []
    window_id = 0

    for record in trial_records:
        n_samples = int(record["n_samples"])
        if n_samples < window_size:
            continue

        max_start = n_samples - window_size
        for start in range(0, max_start + 1, stride_size):
            stop = start + window_size
            window_id += 1
            window_records.append(
                {
                    "window_id": window_id,
                    "subject_id": int(record["subject_id"]),
                    "session_id": int(record["session_id"]),
                    "trial_id": int(record["trial_id"]),
                    "emotion": str(record["emotion"]),
                    "label_id": int(record["label_id"]),
                    "start_sample": start,
                    "stop_sample": stop,
                    "window_size": window_size,
                    "file_path": str(record["file_path"]),
                    "mat_key": str(record["mat_key"]),
                }
            )

    return window_records


def write_csv(path: Path, rows: List[Dict[str, object]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def summarize_records(
    trial_records: List[Dict[str, object]],
    window_records: List[Dict[str, object]],
    sampling_rate: int,
    window_seconds: float,
    stride_seconds: float,
) -> Dict[str, object]:
    trial_counts: Dict[str, int] = {}
    window_counts: Dict[str, int] = {}

    for record in trial_records:
        emotion = str(record["emotion"])
        trial_counts[emotion] = trial_counts.get(emotion, 0) + 1

    for record in window_records:
        emotion = str(record["emotion"])
        window_counts[emotion] = window_counts.get(emotion, 0) + 1

    sample_lengths = [int(record["n_samples"]) for record in trial_records]

    return {
        "num_subjects": len({int(record["subject_id"]) for record in trial_records}),
        "num_trials": len(trial_records),
        "num_windows": len(window_records),
        "sampling_rate_hz": sampling_rate,
        "window_seconds": window_seconds,
        "stride_seconds": stride_seconds,
        "min_samples_per_trial": int(min(sample_lengths)),
        "max_samples_per_trial": int(max(sample_lengths)),
        "trial_count_by_emotion": trial_counts,
        "window_count_by_emotion": window_counts,
        "emotion_to_id": EMOTION_TO_ID,
    }


def read_window_csv(window_csv: Path) -> List[Dict[str, str]]:
    with window_csv.open("r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_window_signal(
    record: Dict[str, str],
    normalize: bool = True,
    eps: float = 1e-6,
) -> Tuple[np.ndarray, int]:
    mat_data = sio.loadmat(record["file_path"])
    signal = np.asarray(mat_data[record["mat_key"]], dtype=np.float32)
    start = int(record["start_sample"])
    stop = int(record["stop_sample"])
    window = signal[:, start:stop]

    if normalize:
        mean = window.mean(axis=1, keepdims=True)
        std = window.std(axis=1, keepdims=True)
        window = (window - mean) / np.maximum(std, eps)

    return window.astype(np.float32), int(record["label_id"])


def main() -> None:
    parser = argparse.ArgumentParser(description="Build SEED-VII EEG trial and window manifests.")
    parser.add_argument("--eeg-dir", type=Path, default=DEFAULT_EEG_DIR)
    parser.add_argument("--label-file", type=Path, default=DEFAULT_LABEL_FILE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--sampling-rate", type=int, default=DEFAULT_SAMPLING_RATE)
    parser.add_argument("--window-seconds", type=float, default=4.0)
    parser.add_argument("--stride-seconds", type=float, default=2.0)
    args = parser.parse_args()

    if not args.eeg_dir.exists():
        raise FileNotFoundError(f"EEG directory not found: {args.eeg_dir}")
    if not args.label_file.exists():
        raise FileNotFoundError(f"Label file not found: {args.label_file}")

    trial_labels = load_trial_labels(args.label_file)
    trial_records = build_trial_records(args.eeg_dir, trial_labels, args.sampling_rate)
    window_records = build_window_records(
        trial_records,
        args.sampling_rate,
        args.window_seconds,
        args.stride_seconds,
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)

    trial_manifest_path = args.output_dir / "trial_manifest.csv"
    window_manifest_path = args.output_dir / "window_manifest.csv"
    summary_path = args.output_dir / "summary.json"

    write_csv(
        trial_manifest_path,
        trial_records,
        [
            "subject_id",
            "session_id",
            "trial_id",
            "emotion",
            "label_id",
            "n_channels",
            "n_samples",
            "duration_sec",
            "file_path",
            "mat_key",
        ],
    )
    write_csv(
        window_manifest_path,
        window_records,
        [
            "window_id",
            "subject_id",
            "session_id",
            "trial_id",
            "emotion",
            "label_id",
            "start_sample",
            "stop_sample",
            "window_size",
            "file_path",
            "mat_key",
        ],
    )

    summary = summarize_records(
        trial_records,
        window_records,
        args.sampling_rate,
        args.window_seconds,
        args.stride_seconds,
    )
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(f"Trial manifest written to: {trial_manifest_path}")
    print(f"Window manifest written to: {window_manifest_path}")
    print(f"Summary written to: {summary_path}")
    print(
        f"Built {len(trial_records)} trials and {len(window_records)} windows "
        f"at {args.window_seconds:.2f}s / {args.stride_seconds:.2f}s."
    )


if __name__ == "__main__":
    main()
